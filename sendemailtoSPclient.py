import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Border, Side, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
import mimetypes
import os
import urllib.parse
from openpyxl.utils.dataframe import dataframe_to_rows
from email.header import make_header

# ================== 配置参数 ==================
pricingdate = pd.to_datetime("2025-05-15")
sender = "fdi_trade_service@xyzq.com.cn"
email_body = """您好，估值文件如附，请查收，如对估值报告有疑问，请联系：fdi_trade_service@xyzq.com.cn，谢谢！
免责声明：

1、兴业证券股份有限公司（简称“兴业证券”）依据商业合理原则，基于我方内部模型计算得出相关衍生品合约于相应日期的估值报告（统称为“估值报告”），该估值报告计算结果仅供贵方参考，并非兴业证券对贵方所持相关衍生品合约权益价值的保证，亦不代表兴业证券对贵方是否应当追加保证金（包括追加预付金等）以及追加保证金金额作出任何意思表示。若贵方在相关衍生品合约下需追加保证金，兴业证券将另行通知贵方。

2、估值报告基于兴业证券内部模型计算得出，该内部模型的调整、优化、参数使用等由兴业证券依据商业合理原则自行决定。

3、估值报告并不构成兴业证券对贵方有关任何事项（包括但不限于相关衍生品合约的提前终止价格、价值以及计算基准、计算公式）的要约或承诺。若贵方拟提前终止相关衍生品合约，应当另行向兴业证券询价。

4、请贵方审慎核对该估值报告。兴业证券不对贵方或任何第三方使用估值报告产生的结果承担任何责任。如对本估值报告有疑问，请及时回复该邮箱进行咨询、反馈。

5、若估值报告因政策法规变化、自然灾害、系统故障、行情传输错误、行情不稳定等不可抗力出现延迟、中断、错误等，兴业证券不承担任何责任。

6、估值报告不应作为投资相关衍生品合约的资产管理产品（以下简称“资管产品”）进行任何份额或权益变更的依据。若资管产品基于估值报告载明的信息进行开放申购、赎回等份额或权益变更操作，兴业证券不对该等操作的合理性、公允性和法律适当性承担任何责任。若贵方与资管产品投资者就估值报告载明的估值产生任何争议，由贵方自行解决，兴业证券不因此对资管产品投资者承担任何责任。

7、贵方不得向除资管产品的外包服务机构、托管机构或投资顾问以外的第三方披露本估值报告。

8.估值报告的计算结果以交易不可发生提前终止为前提，若贵方确有临时提前终止交易需求，应另行向兴业证券询价，估值报告计算结果不作为提前终止报价的参考依据。"""

# ================== 核心逻辑 ==================
# 读取原始数据
modifytrade = pd.read_excel("optionlist.xlsx", sheet_name=7)  # 第8个sheet
customerlist = modifytrade['CustomerName'].unique()



# 计算annualizedPV
modifytrade['Startdate'] = pd.to_datetime(modifytrade['Startdate'])

modifytrade['livingdays'] = (pricingdate - modifytrade['Startdate']).dt.days

conditions = [
    modifytrade['ModifyType'] != 2,
    modifytrade['ModifyType'] == 2
]
choices = [
    modifytrade['Premium'] - modifytrade['NotionalPrincipal'] * modifytrade['Premium_Percentage'] * modifytrade[
        'livingdays'] / 365,
    modifytrade['NotionalPrincipal'] * (1 + modifytrade['Fixedrate_Annual'] * modifytrade['livingdays'] / 365)
]
modifytrade['annualizedPV'] = np.select(conditions, choices)

# 通用样式设置
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal='center', wrap_text=True)

for customer in customerlist:
    customer_data = modifytrade[modifytrade['CustomerName'] == customer]
    #receiver = ["fdi_derivs_trading@xyzq.com.cn"] + customer_data.iloc[0]['MailAddress'].split(',')
    receiver = ["lvjiawei@xyzq.com.cn"]

    # ========== 处理估值报告 ==========
    if customer_data.iloc[0]['ModifyType'] == 2:
        # 收益凭证估值报告处理
        report_name = f"兴业证券-{customer}-收益凭证估值报告-{pricingdate.strftime('%Y-%m-%d')}.xlsx"
        file_path = os.path.join("D:\\workingdirectory\\ValuationFile", report_name)

        wb = openpyxl.load_workbook(file_path)
        ws = wb.worksheets[0]
        start_row = 5
        end_row = start_row + len(customer_data) - 1
        n_col = ws['N']
        i_col = ws['I']
        raw_pv_list = [cell.value for cell in n_col[start_row - 1:end_row]]
        i_values = [cell.value for cell in i_col[start_row - 1:end_row]]
        modified_pv = np.maximum(customer_data['annualizedPV'].values, raw_pv_list)
        modified_nav = np.round(modified_pv / i_values, 4)

        for idx, (pv, nav) in enumerate(zip(modified_pv, modified_nav), start=start_row):
            ws[f'N{idx}'] = pv
            ws[f'O{idx}'] = nav
        wb.save(file_path)
    else:
        # 场外交易估值报告处理（ModifyType=1或3）
        report_name = f"兴业证券-{customer}-场外交易估值报告-{pricingdate.strftime('%Y-%m-%d')}.xlsx"
        file_path = os.path.join("D:\\workingdirectory\\ValuationFile", report_name)
        raw_wb = openpyxl.load_workbook(file_path)
        raw_ws1 = raw_wb.worksheets[0]
        raw_ws2 = raw_wb.worksheets[1]
        start_row = 3
        if customer_data.iloc[0]['ModifyType'] == 1:
            # 加载模板文件
            wb = openpyxl.load_workbook("D:\\workingdirectory\\valuationreport_options_modified.xlsx")
            ws1 = wb.worksheets[0]
            ws2 = wb.worksheets[1]

            # 填充首部信息
            ws1['B1'] = customer
            ws1['B2'] = pricingdate.strftime("%Y/%m/%d")

            # 填充现金摘要（假设原始现金摘要在第6行）
            cash_summary = []
            for cell in raw_ws1[6][0:5]:
                cash_summary.append(cell.value)
            #cash_summary = pd.DataFrame([cash_summary])

            for col_idx, value in enumerate(cash_summary, start=1):
                cell = ws1.cell(row=6, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = center_alignment

            # 填充详细交易数据（核心修改逻辑）
            tradedata = []
            for row in raw_ws1.iter_rows(min_row=3, min_col=1, max_col=19, values_only=True):
                #rowdata = [cell.value for cell in row]
                tradedata.append(row)
            tradedata = pd.DataFrame(tradedata)
            tradedata[19]=tradedata[18]
            tradedata[18]=customer_data['annualizedPV']
            tradedata[18]=tradedata[[17,18]].max(axis=1)


            # 写入工作表并设置样式
            for row_idx, row in enumerate(dataframe_to_rows(tradedata[[*range(20)]], index=False, header=False),
                                          start=start_row):
                ws2.append(row)


            # 保存文件
            wb.save(file_path)
        elif customer_data.iloc[0]['ModifyType'] == 3:
            cells = raw_ws2.iter_rows(min_row=start_row, min_col=18, max_col=18, values_only=False)
            for idx, cell in enumerate(cells):
                current_val = cell[0].value
                if current_val < customer_data['annualizedPV'].tolist()[idx]:
                    cell[0].value = customer_data['annualizedPV'].tolist()[idx]
            raw_wb.save(file_path)



    # ========== 发送邮件 ==========
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = ", ".join(receiver)
    msg['Subject'] = report_name.split('.')[0]
    msg.attach(MIMEText(email_body, 'plain', 'utf-8'))

    with open(file_path, 'rb') as f:
        full_filename = os.path.basename(file_path)  # 带扩展名的文件名
        #mime_type, _ = mimetypes.guess_type(full_filename)
        #if not mime_type:
           # mime_type = 'application/octet-stream'
        part = MIMEApplication(f.read(), Name=report_name)
        encoded_filename = urllib.parse.quote(report_name)
        filename_header = f"utf-8''{encoded_filename}"
        part['Content-Disposition'] = f'attachment; filename="{report_name}"; filename*={filename_header}'
        part.add_header('Content-Disposition', 'attachment',
                              filename=make_header([(full_filename, 'utf-8')]).encode('utf-8'))
        #part.add_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        #                name=encoded_filename)
        #part["Content-Type"] = 'application/octet-stream'
        #part.add_header('Content-Disposition', 'attachment',
        #                filename=make_header([(report_name, 'utf-8')]).encode('utf-8'))
        #encoded_filename = urllib.parse.quote(full_filename)
        #safe_filename = urllib.parse.quote(full_filename, safe='')
        #part['Content-Disposition'] = (
        #    f'attachment; '
         #   f'filename="{safe_filename}"; '
         #   f'filename*=utf-8\'\'{encoded_filename}'
        #)
        #part.add_header('Content-Type', mime_type, name=encoded_filename)
        #part['Content-Disposition'] = f'attachment; filename="{report_name}"'
        msg.attach(part)

    try:
        # 使用SMTP_SSL加密连接
        with smtplib.SMTP_SSL("mail.xyzq.com.cn", 994) as server:
            server.login(sender, "BkgChK4BgKkXnKre")  # 密码建议从环境变量读取
            server.send_message(msg)
            print(f"邮件发送成功：{customer}")
    except Exception as e:
        print(f"邮件发送失败（{customer}）：{str(e)}")


